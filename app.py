
# -*- coding: utf-8 -*-
"""
Created on Wed Dec 18 10:49:58 2024

@author: sophi

source documentation : https://dash-bootstrap-components.opensource.faculty.ai/docs/components/accordion/
http://127.0.0.1:8050
"""
#cd C:\Users\sophi\myCloud\Sophia\Thesis\Model\Jucar_model\Adrià\App_Jucar
#streamlit run Jucar_river_basin.py
#git save

# Import libraries and modules
import pandas as pd
import numpy as np
import openpyxl
import pysd
import dash
from dash import Dash, dcc, html, Input, Output, State
import dash_bootstrap_components as dbc
import plotly.graph_objs as go

# Precompute initial graph data
workbook = openpyxl.load_workbook("data_initial.xlsx")
workbook.save("data.xlsx")
vensim_model = pysd.load('WEFE Jucar (Simple).py')
initial_qecolAlar_value = 5.18
qecolAlar_value = initial_qecolAlar_value
years_sim = 10
months = np.arange(1, (12*years_sim)+1)
variables_model_initial = vensim_model.run(params={'INITIAL TIME': 1, 'FINAL TIME': 12*years_sim, 'TIME STEP': 1})

initial_outflow = variables_model_initial['Sal Jucar']
initial_deficit = variables_model_initial['DéfQecolAlar']

# Initialize the Dash app
app = Dash(__name__, external_stylesheets=[dbc.themes.FLATLY], suppress_callback_exceptions=True)
app.title = "Júcar River Basin Water Management"

# Define reusable page components
def create_home_page():
    return html.Div([
        html.H1("Júcar River Basin Management Tool", className="text-center mt-0"),
        html.P("Welcome to the Júcar River Basin System Dynamics Model. This tool provides an interactive platform "
               "to analyze and simulate the behavior of the Júcar River Basin under various scenarios."),
        html.H2("Content", className="mt-4"),
        html.Ul([
            html.Li("1. Model presentation - see how the model is formed"),
            html.Li("2. Explore environmental and reservoir management impacts"),
            html.Li("3. Drought management simulation"),
            html.Li("4. Monitor system performance through metrics"),
        ]),
        html.H2("User Guide", className="mt-4"),
        html.Ol([
            html.Li("Navigate using the menu on the left."),
            html.Li("Adjust variables using sliders or select different scenarios using dropdown menus."),
            html.Li("Run simulations to analyze results."),
        ])
    ])

def create_model_presentation():
    return html.Div([
        html.H1("Júcar River Basin Management Tool", className="text-center mt-0"),
        dbc.Tabs([
            dbc.Tab(label="View 1: SYSTEM NETWORK", children=[
                html.P("Overview of the model"),
                html.Img(src="/assets/View_1.PNG", style={"width": "80%", "margin": "auto", "display": "block"})
            ]),
            dbc.Tab(label="View 2", children=[
                html.P("MANCHA ORIENTAL AQUIFER"),
                html.Img(src="/assets/View_2.PNG", style={"width": "80%", "margin": "auto", "display": "block"})
            ]),
            dbc.Tab(label="View 3", children=[
                html.P("WATER DEMAND, SUPPLY AND DEFICIT"),
                html.Img(src="/assets/View_3.PNG", style={"width": "80%", "margin": "auto", "display": "block"})
            ]),
            dbc.Tab(label="View 4", children=[
                html.P("RESERVOIRS OPERATING RULES"),
                html.Img(src="/assets/View_4.PNG", style={"width": "80%", "margin": "auto", "display": "block"})
            ]),
            dbc.Tab(label="View 5", children=[
                html.P("STATE INDEX"),
                html.Img(src="/assets/View_5.PNG", style={"width": "80%", "margin": "auto", "display": "block"})
            ]),
            dbc.Tab(label="View Crops", children=[
                html.P("CROPS"),
                html.Img(src="/assets/Crops.PNG", style={"width": "80%", "margin": "auto", "display": "block"})
            ]),
        ])
    ])

def create_alarcon_page():
    # choose between slider and dropdown
    select_input_type = html.Div([
         html.P("Select Input Type:", className="text-center ml-2"),
         dbc.RadioItems(id="qecolAlar-selector", options=[{"label": "Constant (Slider)", "value": "slider"},{"label": "Vary Monthly (Dropdown)", "value": "dropdown"},],inline=True,value="slider",className="mb-3")
         ])
    # Slider for constant input
    qecolAlar_slider = html.Div([
        html.P("QEcolAlar: This is the environmental flow downstream of Alarcon’s reservoir.", className="text-center ml-2"),
        dcc.Slider(min=0.0, max=10.0,step=0.1,value=initial_qecolAlar_value,marks={i: str(i) for i in range(0, 11)})],
        id="qecolAlar-slider",
        style={"display": "none", "width": "5%", "margin": "0 auto"} # Initially hidden
        ) 
    
    qecolAlar_dropdown = html.Div([
        html.P("QEcolAlar: This is the environmental flow downstream of Alarcon’s reservoir.", className="text-start ml-2"),
        dbc.Input(type="number", min=0.0, max=10.0,step=0.1,value=initial_qecolAlar_value)],
        id="qecolAlar-dropdown",
        style={"display": "none","width": "5%", "margin": "0 auto"} # Initially hidden
        )

    run_simulation = html.Div([
        dbc.Button("Run Simulation",  color="primary", className="text-start ml-2")],
        id="run-simulation",
        style={"display": "none", "width": "5%", "margin": "0 auto"} # Initially hidden
        ) 
    
    result_section = dbc.Spinner(
        html.Div([
            html.P("DéfQEcolAlar: is the deficit regarding the environmental flow. ", className="text-start ml-2"),
            dcc.Graph(
                id="outflow-graph",
                figure={
                    "data": [go.Scatter(x=months, y=initial_outflow, mode="lines", name="Outflow")],
                    "layout": go.Layout(title="Outflow Over Time", xaxis={"title": "Months"}, yaxis={"title": "hm³"})
                }
            ),
            dcc.Graph(
                id="deficit-graph",
                figure={
                    "data": [go.Scatter(x=months, y=initial_deficit, mode="lines", name="Deficit")],
                    "layout": go.Layout(title="Deficit Over Time (DéfQEcolAlar)", xaxis={"title": "Months"}, yaxis={"title": "hm³"})
                }
            )
        ])
    )

    return html.Div([
        html.H1("Alarcón’s Reservoir", className="text-center mt-0"),
        # collapse,
        html.P("Use the slider and run the simulation."),
        select_input_type,
        qecolAlar_slider,
        qecolAlar_dropdown,
        run_simulation,
        result_section
    ])

# Layout
app.layout = dbc.Container(fluid=True, children=[
    dbc.Row([
        dbc.Col(
            dbc.Card([
                dbc.CardHeader("Menu", className="bg-dark text-white text-center py-3"),
                dbc.ListGroup([
                    dbc.ListGroupItem("Home", href="/", active="exact", className="text-dark"),
                    dbc.ListGroupItem("Model presentation", href="/model", active="exact", className="text-dark"),
                    dbc.ListGroupItem("Alarcón’s Reservoir", href="/alarcon", active="exact", className="text-dark"),
                ]),],
                className="shadow-sm", style={"height": "100vh"}), width=1.5),
        dbc.Col(
            [dcc.Location(id="url"), html.Div(id="page-content", style={"padding": "20px"})],   width=10)
        ])
    ])

# Callbacks
@app.callback(
    Output("page-content", "children"),
    Input("url", "pathname")
)
def update_page(pathname):
    if pathname == "/":
        return create_home_page()
    elif pathname == "/model":
        return create_model_presentation()
    elif pathname == "/alarcon":
        return create_alarcon_page()
    return html.Div("404: Page Not Found")

#selection of slider_value
@app.callback(
    [Output("qecolAlar-slider", "style"),
     Output("qecolAlar-dropdown", "style"),
     Output("run-simulation", "style")],
    Input("qecolAlar-selector", "value")
)
def toggle_input(input_type):
    if input_type == "slider":
        return {"display": "block"}, {"display": "none"}, {"display": "block"}
    elif input_type == "dropdown":
        return {"display": "none"}, {"display": "block"}, {"display": "block"}
    return {"display": "none"}, {"display": "none"}, {"display": "none"}

@app.callback(
    [Output("outflow-graph", "figure"),
     Output("deficit-graph", "figure")],
    [State("qecolAlar-selector", "value"),  # Use slider value as state
    State("qecolAlar-slider", "value"),  # Use slider value as state
    State("qecolAlar-dropdown", "value")],  # Use slider value as state
    [Input("run-simulation", "n_clicks")],  # Trigger only on button click
    prevent_initial_call=True
)
def update_Alarcon_graphs(input_type, slider_value,dropdown_value, n_clicks):
    if n_clicks is None:
        raise dash.exceptions.PreventUpdate  # Prevent callback if no clicks
    """
    # Determine input value
    qecolAlar_value = slider_value if input_type == "slider" else dropdown_value
    workbook = openpyxl.load_workbook("data.xlsx")
    sheet = workbook["Demandas"]
    column_name = "QecolAlar"
    #change only until the max row of the column QecoAlar
    # First, find the index of column "QecolAlar
    column_index = None
    for col in sheet[2]:  # Check header row 2
        if col.value == column_name:
            column_index = col.column  
            break
      #find max index
    # Technic : find the last row not empty of the column
    last_row = 2  
    for row in range(3, sheet.max_row+1): 
        if sheet.cell(row=row, column=column_index).value is not None:
           last_row = row  
    # change the values of column QecoAlar with the value wanted and save it into data.xlsx to keep data_initial intact
    for row in range(3, last_row + 1):
        sheet.cell(row=row, column=column_index).value = qecolAlar_value
    workbook.save("data.xlsx")
    print(qecolAlar_value)
    #rerun the model
    vensim_model = pysd.load('WEFE Jucar (Simple).py')
    variables_model = vensim_model.run(params={'INITIAL TIME': 1, 'FINAL TIME': 12*years_sim, 'TIME STEP': 1})
    updated_outflow = variables_model['Sal Jucar']
    updated_deficit = variables_model['DéfQecolAlar']

    
    outflow_figure = {
        "data": [
            go.Scatter(x=months, y=initial_outflow, mode="lines", name=f"Initial Outflow (QecoAlar = {initial_qecolAlar_value})", line=dict(dash="dot")),
            go.Scatter(x=months, y=updated_outflow, mode="lines", name=f"Updated Outflow (QecoAlar = {qecolAlar_value})")
        ],
        "layout": go.Layout(title="Outflow Over Time", xaxis={"title": "Months"}, yaxis={"title": "hm³"})
    }
    deficit_figure = {
        "data": [go.Scatter(x=months, y=initial_deficit, mode="lines", name=f"initial Deficit (QecoAlar ={initial_qecolAlar_value})",line=dict(dash="dot")),
                go.Scatter(x=months, y=updated_deficit, mode="lines", name=f"Uptated Deficit(QecoAlar ={qecolAlar_value})")],

        "layout": go.Layout(title="Deficit Over Time", xaxis={"title": "Months"}, yaxis={"title": "hm³"})
    }

    return outflow_figure, deficit_figure"""

# Run the app
if __name__ == "__main__":
    app.run_server(debug=True)
