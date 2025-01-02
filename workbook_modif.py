
# -*- coding: utf-8 -*-
"""
Created on Wed Dec 18 10:49:58 2024

@author: sophi

source documentation : https://dash-bootstrap-components.opensource.faculty.ai/docs/components/accordion/
http://127.0.0.1:8050
"""
#cd C:\Users\sophi\myCloud\Sophia\Thesis\Model\Jucar_model\Adrià\App_Jucar
#streamlit run Jucar_river_basin.py
#git save

# Import libraries and modules
import pandas as pd
import numpy as np
import openpyxl
import pysd
import dash
from dash import Dash, dcc, html, Input, Output, State
import dash_bootstrap_components as dbc
import plotly.graph_objs as go

# Precompute initial graph data
workbook = openpyxl.load_workbook("data_initial.xlsx")
workbook.save("data.xlsx")
vensim_model = pysd.load('WEFE Jucar (Simple).py')
years_sim = 10
months = np.arange(1, (12*years_sim)+1)
variables_model_initial = vensim_model.run(params={'INITIAL TIME': 1, 'FINAL TIME': 12*years_sim, 'TIME STEP': 1})
#1.Alarcon initial values
initial_qecolAlar_value = 5.18
qecolAlar_value = initial_qecolAlar_value
initial_outflow = variables_model_initial['Sal Jucar']
initial_deficit = variables_model_initial['DéfQecolAlar']
#1.Population growth initial value
initial_variation_rate = 0.00018728
variation_rate = initial_variation_rate
initial_urban_demand = variables_model_initial['Total Demanda Urbana']

# Initialize the Dash app
app = Dash(__name__, external_stylesheets=[dbc.themes.FLATLY], suppress_callback_exceptions=True)
app.title = "Júcar River Basin Water Management"

# Define reusable page components
def create_home_page():
    return html.Div([
        html.H1("Júcar River Basin Management Tool", className="text-center mt-0"),
        html.P("Welcome to the Júcar River Basin System Dynamics Model. This tool provides an interactive platform "
               "to analyze and simulate the behavior of the Júcar River Basin under various scenarios."),
        html.H2("Content", className="mt-4"),
        html.Ul([
            html.Li("1. Model presentation - see how the model is formed"),
            html.Li("2. Explore environmental and reservoir management impacts"),
            html.Li("3. Drought management simulation"),
            html.Li("4. Monitor system performance through metrics"),
        ]),
        html.H2("User Guide", className="mt-4"),
        html.Ol([
            html.Li("Navigate using the menu on the left."),
            html.Li("Adjust variables using sliders or select different scenarios using dropdown menus."),
            html.Li("Run simulations to analyze results."),
        ])
    ])

def create_model_presentation():
    return html.Div([
        html.H1("Júcar River Basin Management Tool", className="text-center mt-0"),
        dbc.Tabs(
            [
                dbc.Tab(label="View 1: SYSTEM NETWORK", children=[
                    html.P("Overview of the model"),
                    html.Img(src="/assets/View_1.PNG", style={"width": "100%", "margin": "auto", "display": "block"})
                ]),
                dbc.Tab(label="Aquifer", children=[
                    html.P("View 2: MANCHA ORIENTAL AQUIFER"),
                    html.Img(src="/assets/View_2.PNG", style={"width": "100%", "margin": "auto", "display": "block"})
                ]),
                dbc.Tab(label="Water demand, supply and deficit", children=[
                    html.P("View 3: WATER DEMAND, SUPPLY AND DEFICIT"),
                    html.Img(src="/assets/View_3.PNG", style={"width": "100%", "margin": "auto", "display": "block"})
                ]),
                dbc.Tab(label="Reservoirs operating rules", children=[
                    html.P("View 4: RESERVOIRS OPERATING RULES"),
                    html.Img(src="/assets/View_4.PNG", style={"width": "100%", "margin": "auto", "display": "block"})
                ]),
                dbc.Tab(label="State index", children=[
                    html.P("View 5: STATE INDEX"),
                    html.Img(src="/assets/View_5.PNG", style={"width": "100%", "margin": "auto", "display": "block"})
                ]),
                dbc.Tab(label="Crops", children=[
                    html.P("View : CROPS"),
                    html.Img(src="/assets/Crops.PNG", style={"width": "100%", "margin": "auto", "display": "block"})
                ]),
            ]
        )
    ])



def create_parameter_panel_Alarcon():
    return html.Div([
        html.H2("Parameter Settings", className="text-center mt-3"),
        html.P("QEcolAlar: This is the environmental flow downstream of Alarcon’s reservoir.", className="text-center ml-2"),
        html.Div([
            # choose between slider and dropdown
            html.P("Select Input Type:"),
            dbc.RadioItems(
                id="qecolAlar-selector",
                options=[
                    {"label": "Constant value", "value": "option1"},
                    {"label": "Monthly dynamic values", "value": "option2"},
                ],
                inline=True,
                value="option1",# Default selection
                className="mb-3"
            ),
        ], className="p-4 bg-light shadow rounded", style={ "margin-bottom": "20px"}),
        # Option 1 : Dropdown for constant input 
        html.Div([
            html.P("Set a constant Qeco:", className="text-center mt-1"),
            dbc.Input(id="qecolAlar-constant-input",type="number", min=0.0, max=10.0, step=0.1, value=initial_qecolAlar_value,
                      style={"width": "100px", "text-align": "center", "margin": "0 auto"})
        ], id="qecolAlar-constant", style={"text-align": "center", "margin-bottom": "20px"}),

        # Option 2 : Dynamic dropdown for 12 months for dynamic input
        html.Div([
            html.P("Set a Qeco for each month:", className="text-center mt-1"),
            html.Div([
                dbc.Row([
                    dbc.Col([html.Label("January"), dbc.Input(type="number", id="jan-value", min=0.0, max=10.0, step=0.1, value=initial_qecolAlar_value)], width=6),
                    dbc.Col([html.Label("February"), dbc.Input(type="number", id="feb-value", min=0.0, max=10.0, step=0.1, value=initial_qecolAlar_value)], width=6),
                ], className="mb-2"),
                dbc.Row([
                    dbc.Col([html.Label("March"), dbc.Input(type="number", id="mar-value", min=0.0, max=10.0, step=0.1, value=initial_qecolAlar_value)], width=6),
                    dbc.Col([html.Label("April"), dbc.Input(type="number", id="apr-value", min=0.0, max=10.0, step=0.1, value=initial_qecolAlar_value)], width=6),
                ], className="mb-2"),
                dbc.Row([
                    dbc.Col([html.Label("May"), dbc.Input(type="number", id="may-value",min=0.0, max=10.0, step=0.1, value=initial_qecolAlar_value)], width=6),
                    dbc.Col([html.Label("June"), dbc.Input(type="number", id="jun-value", min=0.0, max=10.0, step=0.1, value=initial_qecolAlar_value)], width=6),
                ], className="mb-2"),
                dbc.Row([
                    dbc.Col([html.Label("July"), dbc.Input(type="number", id="jul-value", min=0.0, max=10.0, step=0.1, value=initial_qecolAlar_value)], width=6),
                    dbc.Col([html.Label("August"), dbc.Input(type="number", id="aug-value", min=0.0, max=10.0, step=0.1, value=initial_qecolAlar_value)], width=6),
                ], className="mb-2"),
                dbc.Row([
                    dbc.Col([html.Label("September"), dbc.Input(type="number", id="sep-value", min=0.0, max=10.0, step=0.1, value=initial_qecolAlar_value)], width=6),
                    dbc.Col([html.Label("October"), dbc.Input(type="number", id="octo-value",min=0.0, max=10.0, step=0.1, value=initial_qecolAlar_value)], width=6),
                ], className="mb-2"),
                dbc.Row([
                    dbc.Col([html.Label("November"), dbc.Input(type="number", id="nov-value", min=0.0, max=10.0, step=0.1, value=initial_qecolAlar_value)], width=6),
                    dbc.Col([html.Label("December"), dbc.Input(type="number", id="dec-value", min=0.0, max=10.0, step=0.1, value=initial_qecolAlar_value)], width=6),
                ])
            ])
        ], id="qecolAlar-dynamic", 
            style={"display": "none", "margin-bottom": "20px","margin-top": "20px"},
            className="text-center mt-0"),
        #simulation button
        html.Div([
           dbc.Button("Run Simulation", id="run-simulation", color="primary",className="text-center mt-0")
       ],  style={"display": "flex", "justify-content": "center","align-items": "center", "margin-top": "20px",  "margin-bottom": "20px"})
   #], style={"maxWidth": "500px", "margin": "0 auto","display": "flex","flex-direction": "column","align-items": "center"}, className="p-4 bg-light shadow rounded")
    ], style={"maxWidth": "600px", "margin": "20px auto", "padding": "20px", "boxShadow": "0 4px 8px rgba(0,0,0,0.1)"})
def create_parameter_panel_population():
    return html.Div([
        html.H2("Parameter Settings", className="text-center mt-3"),
        html.P("Variation rate: Simulate the population growth for Valencia and Sagunto.", className="text-center ml-2"),
        #Dropdown 
        html.Div([
            html.P("Set Variation rate:", className="text-center mt-1"),
            dbc.Input(id="variation-rate-input",type="number", min=0.000, max=1.000, step=0.0001, value=initial_variation_rate,
                      style={"width": "100px", "text-align": "center", "margin": "0 auto"})
        ], id="variation-rate", style={"text-align": "center", "margin-bottom": "10px"}),
        #simulation button
        html.Div([
           dbc.Button("Run Simulation", id="run-simulation", color="primary",className="text-center mt-0")
       ],  style={"display": "flex", "justify-content": "center","align-items": "center", "margin-top": "20px",  "margin-bottom": "20px"})
   ], style={"maxWidth": "500px", "margin": "0 auto","display": "flex","flex-direction": "column","align-items": "center"}, className="p-4 bg-light shadow rounded")

def create_alarcon_page():
    return dbc.Container(fluid=True,children=[
        html.H1("Alarcón’s Reservoir", className="text-center mt-0"),
        dbc.Row([
            dbc.Col(create_parameter_panel_Alarcon(), width=3,className="bg-light border-right",),
            dbc.Col([
                dbc.Spinner(
                html.Div([
                    html.P("DéfQEcolAlar: Deficit regarding the environmental flow."),
                    dcc.Graph(
                        id="outflow-graph",
                        figure={
                            "data": [go.Scatter(x=months, y=initial_outflow, mode="lines", name="Outflow")],
                            "layout": go.Layout(title="Outflow Over Time", xaxis={"title": "Months"}, yaxis={"title": "hm³"})
                        }),
                    dcc.Graph(
                        id="deficit-graph",
                        figure={
                            "data": [go.Scatter(x=months, y=initial_deficit, mode="lines", name="Deficit")],
                            "layout": go.Layout(title="Deficit Over Time (DéfQEcolAlar)", xaxis={"title": "Months"}, yaxis={"title": "hm³"})
                        }),
                ]))
            ], width=9),
        ])
    ])

def create_population_growth_page():
    return dbc.Container(fluid=True, children=[
        html.H1("Population Growth Analysis", className="text-center mt-0"),
        dbc.Row([
            dbc.Col(create_parameter_panel_population(), width=3, className="bg-light border-right"),
            dbc.Col([
                dbc.Spinner(
                    html.Div([
                        html.P("Population Growth Dynamics: Analyze how population growth impacts urban demand."),
                        dcc.Graph(
                            id="demand-graph",
                            figure={
                                "data": [go.Scatter(x=months, y=initial_urban_demand, mode="lines", name="Total Demanda Urbana")],
                                "layout": go.Layout(title="Total Demanda Urbana", xaxis={"title": "Months"}, yaxis={"title": "hm³"})
                            }
                        ),
                    ])
                )
            ], width=9),
        ])
    ])


# Navigation menu
def create_menu():
    return html.Div(
        dbc.ListGroup(
            [
                dbc.ListGroupItem("Home", href="/", active="exact"),
                dbc.ListGroupItem("Model presentation", href="/model", active="exact"),
                dbc.ListGroupItem("Alarcón’s Reservoir", href="/alarcon", active="exact"),
                dbc.ListGroupItem("Population Growth", href="/population-growth", active="exact"),
            ],
        ),
        style={"width": "250px", "position": "fixed", "left": "0", "top": "0", "background-color": "#f8f9fa", "padding": "45px 10px", "box-shadow": "2px 0 5px rgba(0,0,0,0.1)"}
    )

# Update page routing
@app.callback(
    Output("page-content", "children"),
    Input("url", "pathname")
)
def update_page(pathname):
    if pathname == "/":
        return create_home_page()
    elif pathname == "/model":
        return create_model_presentation()
    elif pathname == "/alarcon":
        return create_alarcon_page()
    elif pathname == "/population-growth":
        return create_population_growth_page()
    return html.Div("404: Page Not Found")
# Layout
app.layout = html.Div([
    dbc.Button("Menu", id="menu-toggle", color="primary", className="mb-2", style={"position": "fixed", "top": "10px", "left": "10px", "zIndex": 1000}),
    dbc.Collapse(create_menu(), id="menu-collapse", is_open=False),
    html.Div([
        dcc.Location(id="url"),
        html.Div(id="page-content", style={"padding": "20px"})
    ], id="main-content", style={"margin-left": "0px", "transition": "margin-left 0.3s ease"})
])


@app.callback(
    [Output("menu-collapse", "is_open"),# Control the is_open state of the Collapse
     Output("main-content", "style")],
    [Input("menu-toggle", "n_clicks")],# Triggered when the toggle button is clicked
    [State("menu-collapse", "is_open")], # Store the current state of the Collapse
)
def toggle_menu(n_clicks, is_open):
    if n_clicks:
        if not is_open:
            return True, {"margin-left": "250px", "transition": "margin-left 0.3s ease"}
        return False, {"margin-left": "0px", "transition": "margin-left 0.3s ease"}
    return is_open, {"margin-left": "0px", "transition": "margin-left 0.3s ease"}


#selection of slider_value
@app.callback(
    [Output("qecolAlar-constant", "style"),
     Output("qecolAlar-dynamic", "style"),
     Output("run-simulation", "style")],
    Input("qecolAlar-selector", "value")
)


def toggle_input(input_type):
    if input_type == "option1":
        # Show dropdown constant, hide dropdown monthly
        return {"display": "block"}, {"display": "none"}, {"display": "block", "margin-bottom": "20px"}
    elif input_type == "option2":
        # Show dropdown, hide slider
        return {"display": "none"}, {"display": "block", "margin-bottom": "20px","margin-top": "20px"} ,{"display": "block", "margin-bottom": "20px"}
    # If no option is selected, hide both
    return {"display": "none"}, {"display": "none"},{"display": "none"}

@app.callback(
    [Output("outflow-graph", "figure"),
      Output("deficit-graph", "figure")],
    [Input("run-simulation", "n_clicks")],  # Button click to trigger
    [State("qecolAlar-selector", "value"),  # Determines input type
     State("qecolAlar-constant-input", "value"),  # Constant value
     State("jan-value", "value"),  # Monthly dynamic values
     State("feb-value", "value"),
     State("mar-value", "value"),
     State("apr-value", "value"),
     State("may-value", "value"),
     State("jun-value", "value"),
     State("jul-value", "value"),
     State("aug-value", "value"),
     State("sep-value", "value"),
     State("octo-value", "value"),
     State("nov-value", "value"),
     State("dec-value", "value")],
    prevent_initial_call=True
)


def update_Alarcon_graphs(n_clicks,input_type, constant_value,jan, feb, mar, apr, may, jun, jul, aug, sep, octo, nov, dec):
    if n_clicks is None:
        raise dash.exceptions.PreventUpdate  # Prevent callback if no clicks
    # Initial monthly vector
    qecolAlar_values = []
    # Determine input value
    if input_type == "option1" :
        qecolAlar_values = [constant_value]*12
    else :
        qecolAlar_values = [jan, feb, mar, apr, may, jun, jul, aug, sep, octo, nov, dec]

    workbook = openpyxl.load_workbook("data.xlsx")
    sheet = workbook["Demandas"]
    column_name = "QecolAlar"
    #change only until the max row of the column QecoAlar
    # First, find the index of column "QecolAlar
    column_index = None
    for col in sheet[2]:  # Check header row 2
        if col.value == column_name:
            column_index = col.column  
            break
      
    # Write the  for years_sim years (years_sim x 12 rows)
    current_row = 3  # Start from row 3 
    for year in range(years_sim):  # Repeat for years_sim years
        for month_index in range(12):  # Write one year (12 months) of data
            sheet.cell(row=current_row, column=column_index).value = qecolAlar_values[month_index]
            current_row += 1  
    workbook.save("data.xlsx")
    #rerun the model
    vensim_model = pysd.load('WEFE Jucar (Simple).py')
    variables_model = vensim_model.run(params={'INITIAL TIME': 1, 'FINAL TIME': 12*years_sim, 'TIME STEP': 1})
    updated_outflow = variables_model['Sal Jucar']
    updated_deficit = variables_model['DéfQecolAlar']

    
    outflow_figure = {
        "data": [
            go.Scatter(x=months, y=initial_outflow, mode="lines", name=f"Initial Outflow (QecoAlar = {initial_qecolAlar_value})", line=dict(dash="dot")),
            go.Scatter(x=months, y=updated_outflow, mode="lines", name="Updated Outflow")
        ],
        "layout": go.Layout(title="Outflow Over Time", xaxis={"title": "Months"}, yaxis={"title": "hm³"})
    }
    deficit_figure = {
        "data": [go.Scatter(x=months, y=initial_deficit, mode="lines", name=f"initial Deficit (QecoAlar ={initial_qecolAlar_value})",line=dict(dash="dot")),
                go.Scatter(x=months, y=updated_deficit, mode="lines", name=f"Uptated Deficit(QecoAlar ={qecolAlar_value})")],

        "layout": go.Layout(title="Deficit Over Time", xaxis={"title": "Months"}, yaxis={"title": "hm³"})
    }

    return outflow_figure, deficit_figure

@app.callback(
    [Output("demand-graph", "figure")],
    [Input("run-simulation", "n_clicks")],  # Button click to trigger
    [State("variation-rate-input", "value")],  
    prevent_initial_call=True
)
def update_population_graphs(n_clicks,variation_rate):
    if n_clicks is None:
        raise dash.exceptions.PreventUpdate  # Prevent callback if no clicks
    
    vensim_model = pysd.load('WEFE Jucar (Simple).py')
    #change Variation and Activate 
    variables_model = vensim_model.run(params={'INITIAL TIME': 1, 'FINAL TIME': 12*years_sim, 'TIME STEP': 1,'Variation Rate': variation_rate,'"Activar/Desactivar"' :1})
    urban_demand = variables_model['Total Demanda Urbana']

    urban_demand_figure = {
        "data": [
            go.Scatter(x=months, y=initial_urban_demand, mode="lines", name=f"Initial Urban demand (Variation rate = {initial_variation_rate})", line=dict(dash="dot")),
            go.Scatter(x=months, y=urban_demand, mode="lines", name=f"Updated Urban demand  (Variation rate = {variation_rate})")
        ],
        "layout": go.Layout(title="Urban demand evolution", xaxis={"title": "Months"}, yaxis={"title": "hm³"})
    }
    return [urban_demand_figure]

# Run the app
if __name__ == "__main__":
    app.run_server(debug=True)
