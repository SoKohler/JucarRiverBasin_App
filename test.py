# -*- coding: utf-8 -*-
"""
Created on Wed Dec 18 10:49:58 2024

@author: sophi

source documentation: https://dash-bootstrap-components.opensource.faculty.ai/docs/components/accordion/
http://127.0.0.1:8050
"""
# Import libraries and modules
import pandas as pd
import numpy as np
import openpyxl
import pysd
import dash
from dash import Dash, dcc, html, Input, Output, State
import dash_bootstrap_components as dbc
import plotly.graph_objs as go

# Precompute initial graph data
workbook = openpyxl.load_workbook("data_initial.xlsx")
workbook.save("data.xlsx")
vensim_model = pysd.load('WEFE Jucar (Simple).py')
initial_qecolAlar_value = 5.18
qecolAlar_value = initial_qecolAlar_value
years_sim = 10
months = np.arange(1, (12 * years_sim) + 1)
variables_model_initial = vensim_model.run(params={'INITIAL TIME': 1, 'FINAL TIME': 12 * years_sim, 'TIME STEP': 1})

initial_outflow = variables_model_initial['Sal Jucar']
initial_deficit = variables_model_initial['DéfQecolAlar']

# Initialize the Dash app
app = Dash(__name__, external_stylesheets=[dbc.themes.FLATLY], suppress_callback_exceptions=True)
app.title = "Júcar River Basin Water Management"

# Define reusable page components
def create_home_page():
    return html.Div([
        html.H1("Júcar River Basin Management Tool", className="text-center mt-0"),
        html.P("Welcome to the Júcar River Basin System Dynamics Model. This tool provides an interactive platform "
               "to analyze and simulate the behavior of the Júcar River Basin under various scenarios."),
        html.H2("Content", className="mt-4"),
        html.Ul([
            html.Li("1. Model presentation - see how the model is formed"),
            html.Li("2. Explore environmental and reservoir management impacts"),
            html.Li("3. Drought management simulation"),
            html.Li("4. Monitor system performance through metrics"),
        ]),
        html.H2("User Guide", className="mt-4"),
        html.Ol([
            html.Li("Navigate using the menu on the left."),
            html.Li("Adjust variables using sliders or select different scenarios using dropdown menus."),
            html.Li("Run simulations to analyze results."),
        ])
    ])

def create_model_presentation():
    return html.Div([
        html.H1("Júcar River Basin Management Tool", className="text-center mt-0"),
        dbc.Tabs([
            dbc.Tab(label="View 1: SYSTEM NETWORK", children=[
                html.P("Overview of the model"),
                html.Img(src="/assets/View_1.PNG", style={"width": "80%", "margin": "auto", "display": "block"})
            ]),
            dbc.Tab(label="View 2", children=[
                html.P("MANCHA ORIENTAL AQUIFER"),
                html.Img(src="/assets/View_2.PNG", style={"width": "80%", "margin": "auto", "display": "block"})
            ]),
            dbc.Tab(label="View 3", children=[
                html.P("WATER DEMAND, SUPPLY AND DEFICIT"),
                html.Img(src="/assets/View_3.PNG", style={"width": "80%", "margin": "auto", "display": "block"})
            ]),
            dbc.Tab(label="View 4", children=[
                html.P("RESERVOIRS OPERATING RULES"),
                html.Img(src="/assets/View_4.PNG", style={"width": "80%", "margin": "auto", "display": "block"})
            ]),
            dbc.Tab(label="View 5", children=[
                html.P("STATE INDEX"),
                html.Img(src="/assets/View_5.PNG", style={"width": "80%", "margin": "auto", "display": "block"})
            ]),
            dbc.Tab(label="View Crops", children=[
                html.P("CROPS"),
                html.Img(src="/assets/Crops.PNG", style={"width": "80%", "margin": "auto", "display": "block"})
            ]),
        ])
    ])

def create_parameter_panel():
    return html.Div([
        html.H2("Parameter Settings", className="text-center mt-3"),
        html.Div([
            html.P("QEcolAlar: This is the environmental flow downstream of Alarcon’s reservoir.", className="text-center ml-2"),
            dcc.Slider(min=0.0, max=10.0, step=0.1, value=initial_qecolAlar_value, marks={i: str(i) for i in range(0, 11, 5)},
                       id="qecolAlar-slider")
        ]),
        html.Div([
            dbc.Button("Run Simulation", color="primary", className="text-start ml-2", id="run-simulation"),
        ], className="p-4 bg-light shadow rounded")
    ], className="p-4 bg-light shadow rounded")


def create_alarcon_page():
    return dbc.Container(fluid=True, children=[
        html.H1("Alarcón’s Reservoir", className="text-center mt-0"),
        dbc.Row([
            dbc.Col(create_parameter_panel(), width=3, className="bg-light border-right"),
            dbc.Col([
                dbc.Spinner(
                    html.Div([
                        html.P("DéfQEcolAlar: Deficit regarding the environmental flow."),
                        dcc.Graph(
                            id="outflow-graph",
                            figure={
                                "data": [go.Scatter(x=months, y=initial_outflow, mode="lines", name="Outflow")],
                                "layout": go.Layout(title="Outflow Over Time", xaxis={"title": "Months"}, yaxis={"title": "hm³"})
                            }
                        ),
                        dcc.Graph(
                            id="deficit-graph",
                            figure={
                                "data": [go.Scatter(x=months, y=initial_deficit, mode="lines", name="Deficit")],
                                "layout": go.Layout(title="Deficit Over Time (DéfQEcolAlar)", xaxis={"title": "Months"}, yaxis={"title": "hm³"})
                            }
                        ),
                    ])
                )
            ], width=9),
        ])
    ])


# Layout
app.layout = dbc.Container(fluid=True, children=[
    dbc.Row([
        dbc.Col(
            html.Div([
                dbc.Button("Menu", id="menu-toggle", color="primary", className="mb-0 w-50"),
                dbc.Collapse(
                    dbc.Card([
                        dbc.ListGroup([
                            dbc.ListGroupItem("Home", href="/", active="exact", className="text-white"),
                            dbc.ListGroupItem("Model presentation", href="/model", active="exact", className="text-white"),
                            dbc.ListGroupItem("Alarcón’s Reservoir", href="/alarcon", active="exact", className="text-white"),
                        ]),
                    ], className="shadow-sm"),
                    id="menu-collapse",
                    is_open=False
                )
            ]),
            width=2
        ),
        dbc.Col(
            [dcc.Location(id="url"), html.Div(id="page-content", style={"padding": "20px"})],
            width=10
        )
    ])
])

# Callbacks
@app.callback(
    Output("menu-collapse", "is_open"),
    Input("menu-toggle", "n_clicks"),
    State("menu-collapse", "is_open")
)
def toggle_menu(n_clicks, is_open):
    if n_clicks:
        return not is_open
    return is_open

@app.callback(
    Output("page-content", "children"),
    Input("url", "pathname")
)
def update_page(pathname):
    if pathname == "/":
        return create_home_page()
    elif pathname == "/model":
        return create_model_presentation()
    elif pathname == "/alarcon":
        return create_alarcon_page()
    return html.Div("404: Page Not Found")


@app.callback(
    [Output("outflow-graph", "figure"),
     Output("deficit-graph", "figure")],
    [State("qecolAlar-slider", "value")],
    [Input("run-simulation", "n_clicks")],
    prevent_initial_call=True
)
def update_Alarcon_graphs(slider_value, n_clicks):
    # Check if the simulation button is clicked
    if n_clicks is None:
        raise dash.exceptions.PreventUpdate

    # Debugging: Check if slider value is None or if it is properly passed
    print(f"Slider Value: {slider_value}")

    if slider_value is None:
        print("Error: Slider value is None!")
        return dash.no_update, dash.no_update

    qecolAlar_value = slider_value  # Get the value from the slider

    # Debugging: Check if qecolAlar_value is correctly set
    print(f"qecolAlar_value: {qecolAlar_value}")

    # Load the data and update the QecolAlar value in the Excel file
    workbook = openpyxl.load_workbook("data.xlsx")
    sheet = workbook["Demandas"]
    column_name = "QecolAlar"

    # Find the column index for QecolAlar
    column_index = None
    for col in sheet[2]:  # Check header row 2
        if col.value == column_name:
            column_index = col.column
            break

    if column_index is None:
        print("Error: QecolAlar column not found!")
        return dash.no_update, dash.no_update

    # Find the last row to update the values
    last_row = 2
    for row in range(3, sheet.max_row + 1):
        if sheet.cell(row=row, column=column_index).value is not None:
            last_row = row

    # Update the QecolAlar values in the column
    for row in range(3, last_row + 1):
        sheet.cell(row=row, column=column_index).value = qecolAlar_value
    workbook.save("data.xlsx")

    # Re-run the model simulation with updated parameters
    vensim_model = pysd.load('WEFE Jucar (Simple).py')
    variables_model = vensim_model.run(params={'INITIAL TIME': 1, 'FINAL TIME': 12 * years_sim, 'TIME STEP': 1})

    # Get the updated outflow and deficit values
    updated_outflow = variables_model['Sal Jucar']
    updated_deficit = variables_model['DéfQecolAlar']

    # Create the figures for outflow and deficit
    outflow_figure = {
        "data": [
            go.Scatter(x=months, y=initial_outflow, mode="lines", name=f"Initial Outflow (QecoAlar = {initial_qecolAlar_value})", line=dict(dash="dot")),
            go.Scatter(x=months, y=updated_outflow, mode="lines", name=f"Updated Outflow (QecoAlar = {qecolAlar_value})")
        ],
        "layout": go.Layout(title="Outflow Over Time", xaxis={"title": "Months"}, yaxis={"title": "hm³"})
    }

    deficit_figure = {
        "data": [
            go.Scatter(x=months, y=initial_deficit, mode="lines", name=f"Initial Deficit (QecoAlar = {initial_qecolAlar_value})", line=dict(dash="dot")),
            go.Scatter(x=months, y=updated_deficit, mode="lines", name=f"Updated Deficit (QecoAlar = {qecolAlar_value})")
        ],
        "layout": go.Layout(title="Deficit Over Time", xaxis={"title": "Months"}, yaxis={"title": "hm³"})
    }

    return outflow_figure, deficit_figure

# Run the app
if __name__ == "__main__":
    app.run_server(debug=True)
